#! /usr/bin/python3.7

import openpyxl

# Create a Workbook Object
wb = openpyxl.Workbook()  # create a Workbook Object
print(wb.sheetnames)  # on creation, Workbook Objects contain a blank sheet called 'Sheet'

# Change Sheet title
ws = wb['Sheet']
ws.title = 'new_name'
print(wb.sheetnames)

# Calling the save() method will save the changes
wb.save('write_example.xlsx')

# Creating sheets
wb.create_sheet(index=0, title='First Sheet')
print(wb.sheetnames)

# Removing sheets
wb.remove(wb['First Sheet'])
print(wb.sheetnames)

# Writing values to cell
ws = wb['new_name']
ws['A1'] = 'Hello World'

# And finally, save.
wb.save('write_example.xlsx')


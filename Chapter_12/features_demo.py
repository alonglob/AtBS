#! /usr/bin/python3.7
# Font objects
# Formulas
# Setting row/column height/width
# Merge and unMerge cells
# Charts

import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()
ws = wb.active
cell_obj = ws['A1']

# Creating a Font Object:
italic24_font = Font(size=24, italic=True)
cell_obj.font = italic24_font

# Formulas:
# passed as a string representing the visual basic command:
ws['C1'] = '=SUM(A1:B1)'

# Setting Row height and Column width:
ws.row_dimensions[1].height = 70
ws.column_dimensions['A'].width = 20

# Merging and unMerging cells:
ws.merge_cells('A1:C1')
ws.unmerge_cells('A1:C1')

# Charts
# important modules
from openpyxl.chart import ScatterChart, Reference, Series

# Create data
wb = openpyxl.Workbook()
ws = wb.active

rows = [
    ['Size', 'Batch 1', 'Batch 2'],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 25],
    [6, 25, 35],
    [7, 20, 40]]

for row in rows:
    ws.append(row)

chart = ScatterChart()
chart.title = 'Scatter Chart'
chart.x_axis.title = 'Size'
chart.y_axis.title = 'percentage'

xvalues = Reference(ws, min_col=1, min_row=2, max_row=7)
for i in range(2, 4):
    values = Reference(ws, min_col=i, min_row=1, max_row=7)
    series = Series(values, xvalues, title_from_data=True)
    chart.series.append(series)

ws.add_chart(chart, "A10")

wb.save("scatter.xlsx")
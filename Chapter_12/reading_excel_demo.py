#! /usr/bin/python3.7

import openpyxl

# Create a workbook Object from a workbook.xlsx file
wb = openpyxl.load_workbook('xlsx/example.xlsx')

# Getting Sheets from a workbook Object
print(wb.sheetnames)  # prints the list of sheets by name
# => ['Sheet1', 'Sheet2', 'Sheet3']
sheet1_object = wb['Sheet1']  # returns a sheet Object of the first sheet in the workbook
print(sheet1_object.title)  # prints the title of the sheet, 'Sheet1'
# => 'Sheet1'
active_sheet = wb.active  # returns a sheet Object of the active sheet
print(active_sheet.title)
# => 'Sheet1' (probably)


# Getting Cells from a Sheet Object
sheet = wb.active
print(sheet['A1'])
# => <Cell 'Sheet1'.A1>
print(sheet['A1'].value)
#=> 2015-04-05 13:34:02
c = sheet['B1']
c = sheet.cell(row=1, column=2)  # also works: sheet.cell(1, 1)
print('Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value)
print('Cell ' + c.coordinate + ' contains ' + c.value)

# Getting sheet size
print(sheet.dimensions)  # returns A1:C7
print(sheet.max_row)  # returns 7
print(sheet.max_column)  # returns 3

# Convert Column Letters to Numbers:
from openpyxl.utils import get_column_letter, column_index_from_string
print(get_column_letter(90))  # 'CL'
print(column_index_from_string('AA'))  # 27

# Getting multiple cells from sheets:
# the first loop returns a row, looping over columns, and the second loops over the row.
for row_of_cell_obj in sheet['A1':'C3']:
    for cell_obj in row_of_cell_obj:
        print(cell_obj.coordinate + ': ' + str(cell_obj.value))
    print('end of row')

# another option for looping over the column:

for column in sheet.columns:  # generates tuples of columns from left to right as
    for cell_obj in column:  # iterates over each column
        print(cell_obj.coordinate + ': ' + str(cell_obj.value))

for row in sheet.rows:  # generates tuples of rows from top to bottom
    for cell_obj in row:  # iterates through the cells in each row
        print(cell_obj.coordinate + ': ' + str(cell_obj.value))

